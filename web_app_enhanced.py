#!/usr/bin/env python3
"""
Enhanced Flask web application for Excel topic classification.
Provides a web interface with file validation, column selection, and real-time progress.

PERFORMANCE OPTIMIZATIONS:
- Lazy loading: Never load full DataFrame until classification starts
- Lightweight metadata: Use openpyxl/csv to get columns without pandas
- Bottom-up trimming: Scan from end to find last non-empty row efficiently
- No DataFrame caching during browse: Only store file path
"""
from __future__ import annotations

import csv
import io
import os
import sys
import webbrowser
import subprocess
import platform
from pathlib import Path
from typing import Tuple, List, Optional, Any
from threading import Timer
from datetime import datetime
from queue import Queue
import threading

from flask import Flask, render_template, request, send_file, jsonify, session
from werkzeug.utils import secure_filename
import pandas as pd

# Import our classification module
from classify_topics import OpenAIConfig, update_topics


app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max file size
app.config['UPLOAD_FOLDER'] = Path(__file__).parent / 'uploads'
app.config['OUTPUT_FOLDER'] = Path(os.path.expanduser('~/Desktop'))  # Save to desktop
app.secret_key = os.urandom(24)  # For session management

# Create necessary directories
app.config['UPLOAD_FOLDER'].mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}

# Store for progress messages per session
progress_queues = {}


def allowed_file(filename: str) -> bool:
    """Check if file has an allowed extension."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def column_letter_to_index(letter: str) -> int:
    """Convert Excel column letter (A, B, AA, etc.) to 0-based index."""
    letter = letter.upper().strip()
    result = 0
    for char in letter:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1


def column_index_to_letter(index: int) -> str:
    """Convert 0-based index to Excel column letter."""
    letter = ''
    index += 1  # Convert to 1-based
    while index > 0:
        index -= 1
        letter = chr(index % 26 + ord('A')) + letter
        index //= 26
    return letter


# ============================================================================
# LIGHTWEIGHT METADATA EXTRACTION (No pandas, instant response)
# ============================================================================

def _get_csv_metadata(filepath: Path) -> Tuple[List[str], int]:
    """
    Get CSV column headers and approximate row count WITHOUT loading into pandas.
    Uses Python's csv module - very fast.
    Returns: (column_names, estimated_row_count)
    """
    with open(filepath, 'r', encoding='utf-8', errors='replace', newline='') as f:
        reader = csv.reader(f)
        columns = next(reader, [])
        
        # Quick row count estimate: count lines without parsing
        # This is O(n) but very fast - just counting newlines
        f.seek(0)
        row_count = sum(1 for _ in f) - 1  # -1 for header
        
    return columns, max(0, row_count)


def _get_excel_metadata(filepath: Path) -> Tuple[List[str], dict]:
    """
    Get Excel sheet names and column headers WITHOUT loading data.
    Uses openpyxl in read_only mode - very fast.
    Returns: (sheet_names, {sheet_name: column_names})
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(filepath, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    
    # Get columns from first sheet only (others loaded on demand)
    first_sheet = wb[sheet_names[0]]
    first_row = next(first_sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    columns = [str(c) if c is not None else f"Column_{i}" for i, c in enumerate(first_row)]
    
    wb.close()
    return sheet_names, {sheet_names[0]: columns}


def _get_excel_sheet_columns(filepath: Path, sheet_name: str) -> List[str]:
    """
    Get column headers for a specific Excel sheet WITHOUT loading data.
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    columns = [str(c) if c is not None else f"Column_{i}" for i, c in enumerate(first_row)]
    wb.close()
    return columns


# ============================================================================
# EFFICIENT DATA LOADING WITH BOTTOM-UP TRIMMING (Only at classification time)
# ============================================================================

def _find_last_non_empty_row_csv(filepath: Path, answer_col_index: int = 0) -> int:
    """
    Find the last row with content in a CSV file.
    Scans bottom-up on the ANSWER COLUMN - finds trailing empty rows correctly.
    Empty rows in the middle are preserved.
    
    Args:
        filepath: Path to CSV file
        answer_col_index: Which column to check (user's answer column)
    
    Returns: Number of data rows (excluding header)
    """
    with open(filepath, 'r', encoding='utf-8', errors='replace', newline='') as f:
        rows = list(csv.reader(f))
    
    if len(rows) <= 1:  # Only header or empty
        return 0
    
    # Scan from bottom up (skip header at index 0)
    for i in range(len(rows) - 1, 0, -1):
        row = rows[i]
        # Check the answer column (not always column 0)
        if row and len(row) > answer_col_index:
            val = row[answer_col_index].strip()
            if val not in ('', 'NaN', 'nan', 'None', 'null'):
                return i  # Return row count (i = number of data rows since header is 0)
    
    return 0


def _find_last_non_empty_row_excel(filepath: Path, sheet_name: str, answer_col_index: int = 0) -> int:
    """
    Find the last row with content in an Excel sheet.
    Scans bottom-up on the ANSWER COLUMN - finds trailing empty rows correctly.
    Empty rows in the middle are preserved.
    
    OPTIMIZED: Reads only the answer column using pandas (usecols=[answer_col_index]).
    Pandas is highly optimized for single column reads.
    Then scans backwards in memory - this correctly handles gaps.
    
    Args:
        filepath: Path to Excel file
        sheet_name: Name of sheet to read
        answer_col_index: Which column to check (user's answer column)
    
    Returns: Number of data rows (excluding header)
    """
    # Read only the answer column - pandas is fast for single column reads
    df_col = pd.read_excel(filepath, sheet_name=sheet_name, usecols=[answer_col_index], dtype=str)
    
    if len(df_col) == 0:
        return 0
    
    # Get answer column values as a list
    answer_col = df_col.iloc[:, 0].tolist()
    
    # Scan backwards to find last non-empty row in the answer column
    # This correctly handles empty rows in the middle - we only trim TRAILING empty rows
    for i in range(len(answer_col) - 1, -1, -1):
        val = answer_col[i]
        if val is not None and not pd.isna(val):
            if isinstance(val, str):
                if val.strip() not in ('', 'NaN', 'nan', 'None', 'null'):
                    return i + 1  # +1 because we want count of rows (1-based)
            else:
                return i + 1
    
    return 0


def load_and_trim_for_classification(
    filepath: Path, 
    sheet_name: Optional[str] = None,
    progress_callback=None,
    answer_col_index: int = 0
) -> pd.DataFrame:
    """
    Load file and trim trailing empty rows EFFICIENTLY.
    This is called only when classification starts.
    
    Uses the ANSWER COLUMN to determine trailing empty rows.
    Empty rows in the middle of the data are preserved.
    
    - For CSV: Finds last non-empty row first, then loads only that many rows
    - For Excel: Same approach with pandas single-column read
    
    Args:
        filepath: Path to file
        sheet_name: Sheet name for Excel files
        progress_callback: Function to report progress
        answer_col_index: Which column contains answers (for trimming check)
    
    Returns: Trimmed DataFrame ready for classification
    """
    extension = filepath.suffix.lower()
    
    if extension == '.csv':
        if progress_callback:
            progress_callback("Scanning for data rows...")
        
        last_row = _find_last_non_empty_row_csv(filepath, answer_col_index)
        
        if progress_callback:
            progress_callback(f"Found {last_row} rows with data, loading...")
        
        # Load only the rows we need
        if last_row > 0:
            df = pd.read_csv(filepath, nrows=last_row, dtype=str, low_memory=False)
        else:
            df = pd.read_csv(filepath, nrows=1, dtype=str, low_memory=False)
            
    elif extension in ('.xlsx', '.xls'):
        if progress_callback:
            progress_callback(f"Scanning sheet '{sheet_name}' for data rows...")
        
        last_row = _find_last_non_empty_row_excel(filepath, sheet_name, answer_col_index)
        
        if progress_callback:
            progress_callback(f"Found {last_row} rows with data, loading...")
        
        # Load only the rows we need
        if last_row > 0:
            df = pd.read_excel(filepath, sheet_name=sheet_name, nrows=last_row, dtype=str)
        else:
            df = pd.read_excel(filepath, sheet_name=sheet_name, nrows=1, dtype=str)
    else:
        raise ValueError(f"Unsupported file type: {extension}")
    
    # Clean column names
    df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
    
    if progress_callback:
        progress_callback(f"Loaded {len(df)} rows × {len(df.columns)} columns")
    
    return df




class ProgressCapture:
    """Capture print statements for progress display."""
    def __init__(self, session_id: str, queue: Queue):
        self.session_id = session_id
        self.queue = queue
        self.terminal = sys.stdout
        
    def write(self, message):
        if message.strip():  # Ignore empty messages
            try:
                self.queue.put(message.strip())
            except Exception:
                pass  # Silently fail if queue is closed
        self.terminal.write(message)
    
    def flush(self):
        self.terminal.flush()


@app.route('/')
def index():
    """Main upload page."""
    return render_template('index_enhanced.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    """
    Handle file upload - LIGHTWEIGHT, NO DATAFRAME LOADING.
    
    Only extracts metadata (columns, sheets) using fast low-level parsing.
    Full DataFrame loading happens only at classification time.
    """
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type. Please upload .xlsx, .xls, or .csv file'}), 400
    
    try:
        # Save uploaded file
        filename = secure_filename(file.filename)
        session_id = os.urandom(16).hex()
        session['session_id'] = session_id
        
        input_path = app.config['UPLOAD_FOLDER'] / f"{session_id}_{filename}"
        file.save(input_path)
        
        extension = input_path.suffix.lower()
        
        # LIGHTWEIGHT METADATA EXTRACTION - No pandas, instant response
        if extension == '.csv':
            columns, row_count = _get_csv_metadata(input_path)
            sheet_names = ['Sheet1']
            default_sheet = 'Sheet1'
        else:
            # Excel file
            sheet_names, columns_by_sheet = _get_excel_metadata(input_path)
            default_sheet = sheet_names[0]
            columns = columns_by_sheet.get(default_sheet, [])
            row_count = None  # Not computed for Excel to save time
        
        # Store metadata in session (NO DataFrame caching)
        session['current_sheet'] = default_sheet
        
        # Build column info for frontend
        column_info = [
            {
                'index': i,
                'letter': column_index_to_letter(i),
                'name': str(col)
            }
            for i, col in enumerate(columns)
        ]
        
        # Store file info in session
        session['uploaded_file'] = str(input_path)
        session['original_filename'] = filename
        session['sheet_names'] = sheet_names
        
        return jsonify({
            'success': True,
            'session_id': session_id,
            'filename': filename,
            'sheets': sheet_names,
            'default_sheet': default_sheet,
            'columns': column_info,
            'num_rows': row_count,  # May be None for Excel (computed at classification time)
            'num_columns': len(columns)
        })
    
    except Exception as e:
        return jsonify({'error': f'Error processing file: {str(e)}'}), 500


@app.route('/load_sheet', methods=['POST'])
def load_sheet():
    """
    Switch to a different Excel sheet - LIGHTWEIGHT, NO DATAFRAME LOADING.
    
    Only extracts column headers using openpyxl.
    Full DataFrame loading happens only at classification time.
    """
    try:
        data = request.json
        sheet_name = data.get('sheet_name')
        
        if 'session_id' not in session or 'uploaded_file' not in session:
            return jsonify({'error': 'No file uploaded'}), 400
        
        filepath = Path(session['uploaded_file'])
        
        if not filepath.exists():
            return jsonify({'error': 'Uploaded file not found'}), 400
        
        # LIGHTWEIGHT: Get only column headers for this sheet
        columns = _get_excel_sheet_columns(filepath, sheet_name)
        
        # Update current sheet in session (NO DataFrame caching)
        session['current_sheet'] = sheet_name
        
        # Build column info for frontend
        column_info = [
            {
                'index': i,
                'letter': column_index_to_letter(i),
                'name': str(col)
            }
            for i, col in enumerate(columns)
        ]
        
        return jsonify({
            'success': True,
            'sheet_name': sheet_name,
            'columns': column_info,
            'num_rows': None,  # Not computed to save time - happens at classification
            'num_columns': len(columns)
        })
    
    except Exception as e:
        return jsonify({'error': f'Error loading sheet: {str(e)}'}), 500


@app.route('/preview', methods=['POST'])
def preview_data():
    """
    Preview data from selected sheet and columns.
    
    LIGHTWEIGHT: Only loads first 5 rows for preview, not the entire file.
    """
    try:
        data = request.json
        sheet_name = data.get('sheet_name')
        answer_col = data.get('answer_column')
        topic_start = data.get('topic_start_column')
        topic_end = data.get('topic_end_column')
        
        if 'uploaded_file' not in session:
            return jsonify({'error': 'No file uploaded'}), 400
        
        filepath = Path(session['uploaded_file'])
        current_sheet = session.get('current_sheet')
        
        if not filepath.exists():
            return jsonify({'error': 'File not found'}), 400
        
        extension = filepath.suffix.lower()
        
        # LIGHTWEIGHT: Load only first 5 rows for preview
        if extension == '.csv':
            df_preview = pd.read_csv(filepath, nrows=5, dtype=str, low_memory=False)
        else:
            df_preview = pd.read_excel(filepath, sheet_name=current_sheet, nrows=5, dtype=str)
        
        # Clean column names
        df_preview.columns = [c.strip() if isinstance(c, str) else c for c in df_preview.columns]
        
        # Convert column letters to indices
        answer_idx = column_letter_to_index(answer_col)
        topic_start_idx = column_letter_to_index(topic_start)
        topic_end_idx = column_letter_to_index(topic_end)
        
        # Validate indices
        if answer_idx >= len(df_preview.columns) or topic_start_idx >= len(df_preview.columns) or topic_end_idx >= len(df_preview.columns):
            return jsonify({'error': 'Invalid column selection'}), 400
        
        if topic_end_idx < topic_start_idx:
            return jsonify({'error': 'End column must be after or equal to start column'}), 400
        
        # Get column names
        answer_column_name = df_preview.columns[answer_idx]
        topic_columns = df_preview.columns[topic_start_idx:topic_end_idx + 1].tolist()
        
        # Get preview data (first 5 rows)
        preview_data = []
        for i in range(len(df_preview)):
            answer_text = str(df_preview.iloc[i, answer_idx])
            row_data = {
                'row_num': i + 1,
                'answer': answer_text[:100] + '...' if len(answer_text) > 100 else answer_text
            }
            preview_data.append(row_data)
        
        # Format topic columns as numbered list
        topic_columns_formatted = [f"{idx}) {topic}" for idx, topic in enumerate(topic_columns, 1)]
        
        return jsonify({
            'success': True,
            'answer_column': answer_column_name,
            'topic_columns': topic_columns,
            'topic_columns_formatted': topic_columns_formatted,
            'num_topics': len(topic_columns),
            'preview': preview_data
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/classify', methods=['POST'])
def classify():
    """
    Start classification process.
    
    This is where the FULL DATA LOADING happens - uses efficient 
    bottom-up trimming to load only non-empty rows.
    """
    try:
        data = request.json
        sheet_name = data.get('sheet_name')
        answer_col = data.get('answer_column')
        topic_start = data.get('topic_start_column')
        topic_end = data.get('topic_end_column')
        
        if 'uploaded_file' not in session:
            return jsonify({'error': 'No file uploaded'}), 400
        
        session_id = session.get('session_id')
        filepath = Path(session['uploaded_file'])
        current_sheet = session.get('current_sheet')
        
        # Initialize progress queue for this session
        progress_queues[session_id] = Queue()
        queue = progress_queues[session_id]
        
        # Convert column letters to indices
        answer_idx = column_letter_to_index(answer_col)
        topic_start_idx = column_letter_to_index(topic_start)
        topic_end_idx = column_letter_to_index(topic_end)
        
        # Load configuration
        config = OpenAIConfig.from_env()
        original_filename = session['original_filename']
        
        # Run classification in a separate thread
        def run_classification():
            # Redirect both stdout and stderr to capture progress
            old_stdout = sys.stdout
            old_stderr = sys.stderr
            progress_capture = ProgressCapture(session_id, queue)
            sys.stdout = progress_capture
            sys.stderr = progress_capture  # Also capture stderr for progress messages
            
            try:
                queue.put("Starting classification...")
                queue.put(f"Model: {config.model}")
                
                # EFFICIENT LOADING: Uses bottom-up scan on ANSWER COLUMN to find last non-empty row,
                # then loads only the rows we need (not all 1M+ empty rows).
                # Empty rows in the middle are preserved - only trailing empty rows are trimmed.
                df = load_and_trim_for_classification(
                    filepath, 
                    current_sheet,
                    progress_callback=lambda msg: queue.put(msg),
                    answer_col_index=answer_idx  # Use the user's answer column for trimming check
                )
                
                # Get column names
                main_column = df.columns[answer_idx]
                topic_cols = df.columns[topic_start_idx:topic_end_idx + 1].tolist()
                
                queue.put(f"Processing {len(df)} rows with {len(topic_cols)} topics")
                
                # Run classification
                df_classified = update_topics(
                    df,
                    main_column,
                    topic_cols,
                    config,
                    limit=None,
                    skip_existing=False,
                )
                
                # Save to desktop as CSV
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                original_name = Path(original_filename).stem
                output_filename = f"{original_name}_classified_{timestamp}.csv"
                output_path = app.config['OUTPUT_FOLDER'] / output_filename
                
                df_classified.to_csv(output_path, index=False)
                
                queue.put(f"✅ COMPLETE: Saved to {output_path}")
                queue.put("DONE")
                
            except Exception as e:
                import traceback
                error_msg = f"❌ ERROR: {str(e)}\n{traceback.format_exc()}"
                queue.put(error_msg)
                queue.put("ERROR")
            finally:
                sys.stdout = old_stdout
                sys.stderr = old_stderr
        
        # Start classification thread
        thread = threading.Thread(target=run_classification)
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'success': True,
            'message': 'Classification started',
            'session_id': session_id
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/progress/<session_id>')
def get_progress(session_id):
    """Get progress updates for a session."""
    if session_id not in progress_queues:
        return jsonify({'messages': [], 'done': False})
    
    messages = []
    done = False
    error = False
    
    # Get all messages from queue
    queue = progress_queues[session_id]
    while not queue.empty():
        msg = queue.get()
        if msg == "DONE":
            done = True
        elif msg == "ERROR":
            done = True
            error = True
        else:
            messages.append(msg)
    
    return jsonify({
        'messages': messages,
        'done': done,
        'error': error
    })


@app.route('/status')
def status():
    """Check if the server is running and configuration is valid."""
    try:
        config = OpenAIConfig.from_env()
        return jsonify({
            'status': 'ready',
            'model': config.model,
            'api_base_url': config.api_base_url
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': str(e)
        }), 500


def open_browser():
    """Open browser after a short delay. On Windows, specifically use Chrome if available, otherwise use default browser."""
    url = 'http://127.0.0.1:5000'
    
    # On Windows, try to use Chrome specifically
    if platform.system() == 'Windows':
        chrome_paths = [
            r'C:\Program Files\Google\Chrome\Application\chrome.exe',
            r'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe',
            os.path.expanduser(r'~\AppData\Local\Google\Chrome\Application\chrome.exe'),
        ]
        
        chrome_path = None
        for path in chrome_paths:
            if os.path.exists(path):
                chrome_path = path
                break
        
        if chrome_path:
            try:
                subprocess.Popen([chrome_path, url])
                return
            except Exception as e:
                print(f"Warning: Could not open Chrome at {chrome_path}: {e}", file=sys.stderr)
                print("Falling back to default browser...", file=sys.stderr)
                # Fall through to default browser
        else:
            # Chrome not found, use default browser
            print("Chrome not found, using default browser...", file=sys.stderr)
    
    # Default behavior: use system default browser (works on macOS and Linux, or Windows fallback)
    webbrowser.open(url)


if __name__ == '__main__':
    # Open browser automatically (works on both macOS and Windows)
    Timer(1, open_browser).start()
    
    # Run Flask app
    print("=" * 50)
    print("  Topic Classification Web App (Enhanced)")
    print("=" * 50)
    print("Starting server...")
    print("Opening browser at http://127.0.0.1:5000")
    print("Press Ctrl+C to stop the server")
    print("=" * 50)
    print()
    
    try:
        app.run(debug=False, port=5000, host='127.0.0.1', threaded=True)
    except OSError as e:
        if "Address already in use" in str(e) or "address is already in use" in str(e):
            print("\n❌ ERROR: Port 5000 is already in use!")
            print("   Another instance might be running, or another app is using port 5000.")
            print("   Please close it and try again, or edit web_app_enhanced.py to use a different port.")
            sys.exit(1)
        else:
            raise

